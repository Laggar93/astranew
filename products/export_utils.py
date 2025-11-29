import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.utils import timezone
import pytz
from django.utils.html import strip_tags
import re


def clean_html(text):
    """Очищает текст от HTML тегов и лишних пробелов"""
    if not text:
        return ''

    # Удаляем HTML теги
    clean_text = strip_tags(text)

    # Заменяем множественные пробелы и переносы на один пробел
    clean_text = re.sub(r'\s+', ' ', clean_text)

    # Убираем пробелы в начале и конце
    clean_text = clean_text.strip()

    return clean_text


def export_products_to_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    base_headers = [
        'title', 'description', 'keywords', 'subcategories', 'leftovers',
        'pop_models', 'product_title', 'product_description', 'product_price',
        'image', 'instock', 'article', 'brands', 'countries', 'filters', 'seo'
    ]

    all_param_headers = set()
    all_char_headers = set()
    all_slide_headers = set()
    all_file_headers = set()

    for product_obj in queryset:
        params_count = product_obj.product_parameters.count()
        for i in range(1, params_count + 1):
            all_param_headers.add(f'product_parameter_title_{i}')
            all_param_headers.add(f'product_parameter_subtitle_{i}')

        chars_count = product_obj.product_chars.count()
        for i in range(1, chars_count + 1):
            all_char_headers.add(f'product_chars_title_{i}')
            all_char_headers.add(f'product_chars_subtitle_{i}')

        slides_count = product_obj.product_slider.count()
        for i in range(1, slides_count + 1):
            all_slide_headers.add(f'product_slider_image_{i}')
            all_slide_headers.add(f'product_slider_image_alt_{i}')

        files_count = product_obj.product_file.count()
        for i in range(1, files_count + 1):
            all_file_headers.add(f'product_file_{i}')
            all_file_headers.add(f'product_file_title_{i}')

    sorted_param_headers = sorted(all_param_headers, key=lambda x: int(x.split('_')[-1]))
    sorted_char_headers = sorted(all_char_headers, key=lambda x: int(x.split('_')[-1]))
    sorted_slide_headers = sorted(all_slide_headers, key=lambda x: int(x.split('_')[-1]))
    sorted_file_headers = sorted(all_file_headers, key=lambda x: int(x.split('_')[-1]))

    headers = base_headers + sorted_param_headers + sorted_char_headers + sorted_slide_headers + sorted_file_headers

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for product_obj in queryset:
        row_data = {}

        row_data['title'] = product_obj.title or ''
        row_data['description'] = product_obj.description or ''
        row_data['keywords'] = product_obj.keywords or ''
        row_data['subcategories'] = product_obj.subcategories.subcategory_title if product_obj.subcategories else ''
        row_data['leftovers'] = 'TRUE' if product_obj.leftovers else 'FALSE'
        row_data['pop_models'] = 'TRUE' if product_obj.pop_models else 'FALSE'
        row_data['product_title'] = product_obj.product_title or ''

        # Очищаем HTML теги из product_description
        row_data['product_description'] = clean_html(product_obj.product_description) if product_obj.product_description else ''

        row_data['product_price'] = product_obj.product_price or ''
        row_data['image'] = product_obj.image.url if product_obj.image else ''
        row_data['instock'] = 'TRUE' if product_obj.instock else 'FALSE'
        row_data['article'] = product_obj.article
        row_data['brands'] = product_obj.brands.brand_title if product_obj.brands else ''
        row_data['countries'] = product_obj.countries.country_title if product_obj.countries else ''

        filters_list = []
        for filter_param in product_obj.filters.all():
            filters_list.append(f"{filter_param.filters.filter_title}:{filter_param.parameter_title}")
        row_data['filters'] = "; ".join(filters_list)

        # Очищаем HTML теги из seo
        row_data['seo'] = clean_html(product_obj.seo) if product_obj.seo else ''

        parameters = product_obj.product_parameters.all().order_by('order')
        for i, param in enumerate(parameters, 1):
            row_data[f'product_parameter_title_{i}'] = param.title or ''
            row_data[f'product_parameter_subtitle_{i}'] = param.subtitle or ''

        chars = product_obj.product_chars.all().order_by('order')
        for i, char in enumerate(chars, 1):
            row_data[f'product_chars_title_{i}'] = char.title or ''
            # Очищаем HTML теги из характеристик
            row_data[f'product_chars_subtitle_{i}'] = clean_html(char.subtitle) if char.subtitle else ''

        slides = product_obj.product_slider.all().order_by('order')
        for i, slide in enumerate(slides, 1):
            row_data[f'product_slider_image_{i}'] = slide.image.url if slide.image else ''
            row_data[f'product_slider_image_alt_{i}'] = slide.alt or ''

        files = product_obj.product_file.all().order_by('order')
        for i, file_obj in enumerate(files, 1):
            row_data[f'product_file_{i}'] = file_obj.file.url if file_obj.file else ''
            row_data[f'product_file_title_{i}'] = file_obj.name or ''

        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            ws.cell(row=row_num, column=col_num, value=value)

        row_num += 1

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    moscow_tz = pytz.timezone('Europe/Moscow')
    moscow_time = timezone.now().astimezone(moscow_tz)

    filename = f"astra_products_export_{moscow_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response