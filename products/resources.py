import os
import requests
from urllib.parse import urlparse
from django.core.files.base import ContentFile
from django.db import transaction
from django.core.exceptions import ValidationError
from django.contrib import messages
from openpyxl import load_workbook
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.utils import timezone


class ProductExporter:
    def export_products(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Товары"

        headers = [
            'article', 'title', 'description', 'keywords', 'product_title',
            'product_description', 'product_price', 'instock', 'seo',
            'leftovers', 'pop_models', 'subcategories', 'image',
            'brands', 'countries', 'filters'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        row_num = 2
        for product in queryset:
            ws.cell(row=row_num, column=1, value=product.article)
            ws.cell(row=row_num, column=2, value=product.title or '')
            ws.cell(row=row_num, column=3, value=product.description or '')
            ws.cell(row=row_num, column=4, value=product.keywords or '')
            ws.cell(row=row_num, column=5, value=product.product_title or '')
            ws.cell(row=row_num, column=6, value=product.product_description or '')
            ws.cell(row=row_num, column=7, value=product.product_price or 0)
            ws.cell(row=row_num, column=8, value=product.instock or '')
            ws.cell(row=row_num, column=9, value=product.seo or '')
            ws.cell(row=row_num, column=10, value=1 if product.leftovers else 0)
            ws.cell(row=row_num, column=11, value=1 if product.pop_models else 0)

            ws.cell(row=row_num, column=12, value=product.subcategories.subcategory_title if product.subcategories else '')
            ws.cell(row=row_num, column=13, value=product.image.url if product.image else '')
            ws.cell(row=row_num, column=14, value=product.brands.brand_title if product.brands else '')
            ws.cell(row=row_num, column=15, value=product.countries.country_title if product.countries else '')

            filters_list = []
            for filter_param in product.filters.all():
                filters_list.append(f"{filter_param.filters.filter_title}:{filter_param.parameter_title}")
            ws.cell(row=row_num, column=16, value="; ".join(filters_list))

            self.add_product_parameters(ws, product, row_num)

            self.add_product_chars(ws, product, row_num)

            self.add_product_slider(ws, product, row_num)

            self.add_product_files(ws, product, row_num)

            row_num += 1

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        return wb

    def add_product_parameters(self, ws, product, row_num):
        parameters = product.product_parameters_set.all().order_by('order')
        for i, param in enumerate(parameters, 1):
            title_col = 17 + (i - 1) * 2
            subtitle_col = title_col + 1

            if row_num == 2:
                ws.cell(row=1, column=title_col, value=f'product_parameter_title_{i}')
                ws.cell(row=1, column=subtitle_col, value=f'product_parameter_subtitle_{i}')

            ws.cell(row=row_num, column=title_col, value=param.title)
            ws.cell(row=row_num, column=subtitle_col, value=param.subtitle)

    def add_product_chars(self, ws, product, row_num):
        chars = product.product_chars_set.all().order_by('order')
        for i, char in enumerate(chars, 1):
            title_col = 27 + (i - 1) * 2
            subtitle_col = title_col + 1

            if row_num == 2:
                ws.cell(row=1, column=title_col, value=f'product_chars_title_{i}')
                ws.cell(row=1, column=subtitle_col, value=f'product_chars_subtitle_{i}')

            ws.cell(row=row_num, column=title_col, value=char.title)
            ws.cell(row=row_num, column=subtitle_col, value=char.subtitle)

    def add_product_slider(self, ws, product, row_num):
        slides = product.product_slider_set.all().order_by('order')
        for i, slide in enumerate(slides, 1):
            image_col = 47 + (i - 1) * 2
            alt_col = image_col + 1

            if row_num == 2:
                ws.cell(row=1, column=image_col, value=f'product_slider_image_{i}')
                ws.cell(row=1, column=alt_col, value=f'product_slider_image_alt_{i}')

            ws.cell(row=row_num, column=image_col, value=slide.image.url if slide.image else '')
            ws.cell(row=row_num, column=alt_col, value=slide.alt or '')

    def add_product_files(self, ws, product, row_num):
        files = product.product_file_set.all().order_by('order')
        for i, file_obj in enumerate(files, 1):
            file_col = 67 + (i - 1) * 2
            title_col = file_col + 1

            if row_num == 2:
                ws.cell(row=1, column=file_col, value=f'product_file_{i}')
                ws.cell(row=1, column=title_col, value=f'product_file_title_{i}')

            ws.cell(row=row_num, column=file_col, value=file_obj.file.url if file_obj.file else '')
            ws.cell(row=row_num, column=title_col, value=file_obj.name or '')

class ProductImporter:
    def __init__(self, request):
        self.request = request
        self.errors = []
        self.success_count = 0
        self.updated_count = 0
        self.slider_counter = 1

    def import_file(self, file):
        try:
            wb = load_workbook(filename=file, data_only=True)
            sheet = wb.active

            headers = [cell.value for cell in sheet[1]]

            with transaction.atomic():
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        continue

                    try:
                        self.process_row(row, headers, row_num)
                    except Exception as e:
                        self.errors.append(f"Ошибка в строке {row_num}: {str(e)}")
                        continue

            self.create_message()

        except Exception as e:
            self.errors.append(f"Ошибка чтения файла: {str(e)}")
            self.create_message()

    def process_row(self, row, headers, row_num):
        row_data = dict(zip(headers, row))

        article = row_data.get('article')
        if not article:
            raise ValidationError("Поле 'article' обязательно для заполнения")

        product, created = self.get_or_create_product(article, row_data)

        original_image = product.image.name if product.image else None

        self.update_product_fields(product, row_data, row_num)

        self.process_related_data(product, row_data, row_num)

        product.save()

        current_image = product.image.name if product.image else None
        if product.image and current_image != original_image:
            self.create_resizes(product)

        if created:
            self.success_count += 1
        else:
            self.updated_count += 1

    def create_resizes(self, product):
        try:
            from astra.functions import resize_img

            if product.image:
                print(f"Создание ресайзов для продукта {product.article}")

                product.image = resize_img(product.image, product.image, [1280, 960], 'jpeg')

                product.image_detail_png2x = resize_img(
                    product.image_detail_png2x if hasattr(product, 'image_detail_png2x') else None,
                    product.image,
                    [832, 614],
                    'jpeg'
                )
                product.image_detail_png = resize_img(
                    product.image_detail_png if hasattr(product, 'image_detail_png') else None,
                    product.image,
                    [416, 307],
                    'jpeg'
                )
                product.image_detail_webp = resize_img(
                    product.image_detail_webp if hasattr(product, 'image_detail_webp') else None,
                    product.image,
                    [832, 614],
                    'webp'
                )
                product.image_other_png2x = resize_img(
                    product.image_other_png2x if hasattr(product, 'image_other_png2x') else None,
                    product.image,
                    [608, 450],
                    'jpeg'
                )
                product.image_other_png = resize_img(
                    product.image_other_png if hasattr(product, 'image_other_png') else None,
                    product.image,
                    [304, 225],
                    'jpeg'
                )
                product.image_other_webp = resize_img(
                    product.image_other_webp if hasattr(product, 'image_other_webp') else None,
                    product.image,
                    [608, 450],
                    'webp'
                )

                product.save()
                print(f"Ресайзы созданы для продукта {product.article}")

        except Exception as e:
            print(f"Ошибка создания ресайзов: {e}")
            import traceback
            traceback.print_exc()

    def get_or_create_product(self, article, row_data):
        from .models import product

        try:
            product_obj = product.objects.get(article=article)
            created = False
            print(f"Найден существующий продукт: {article}")
        except product.DoesNotExist:
            last_product = product.objects.order_by('-id').first()
            new_id = last_product.id + 1 if last_product else 1

            product_obj = product(
                id=new_id,
                article=article,
                order=new_id
            )
            created = True
            print(f"Создан новый продукт: {article}")

        return product_obj, created

    def update_product_fields(self, product, row_data, row_num):
        basic_fields = ['title', 'description', 'keywords', 'product_title',
                        'product_description', 'product_price', 'instock', 'seo']

        for field in basic_fields:
            value = row_data.get(field)
            if value is not None and value != '':
                setattr(product, field, value)

        bool_fields = ['leftovers', 'pop_models']
        for field in bool_fields:
            value = row_data.get(field)
            if value is not None and value != '':
                setattr(product, field, bool(value))

        subcategory_name = row_data.get('subcategories')
        if subcategory_name:
            product.subcategories = self.get_or_create_subcategory(subcategory_name)

        image_url = row_data.get('image')
        if image_url:
            self.handle_product_image(product, image_url, row_num)

        brand_name = row_data.get('brands')
        if brand_name:
            product.brands = self.get_or_create_brand(brand_name)

        country_name = row_data.get('countries')
        if country_name:
            product.countries = self.get_or_create_country(country_name)

        filters_data = row_data.get('filters')
        if filters_data:
            self.handle_filters(product, filters_data)

    def get_or_create_subcategory(self, subcategory_name):
        from .models import subcategories, categories

        if not subcategory_name:
            return None

        clean_name = subcategory_name.replace('Иконка', '').strip()

        existing = subcategories.objects.filter(subcategory_title=clean_name)

        if existing.exists():
            subcat = existing.first()

            if not subcat.subcategory_description:
                subcat.subcategory_description = f"Раздел {clean_name}"
                subcat.save()
            return subcat

        first_category = categories.objects.first()
        if not first_category:
            raise ValidationError("Нет существующих категорий для создания подкатегории")

        last_subcat = subcategories.objects.order_by('-order').first()
        new_order = last_subcat.order + 1 if last_subcat else 1

        subcategory = subcategories(
            subcategory_title=clean_name,
            categories=first_category,
            order=new_order,
            subcategory_description=f"Раздел {clean_name}"
        )
        subcategory.save()

        return subcategory

    def get_or_create_brand(self, brand_name):
        from .models import brands

        brand, created = brands.objects.get_or_create(brand_title=brand_name)
        return brand

    def get_or_create_country(self, country_name):
        from .models import countries

        country, created = countries.objects.get_or_create(country_title=country_name)
        return country

    def handle_filters(self, product, filters_data):
        from .models import filters, filters_parameters

        if not isinstance(filters_data, str):
            return

        filter_items = [item.strip() for item in filters_data.split(';') if item.strip()]

        for filter_item in filter_items:
            if ':' in filter_item:
                filter_title, parameter_title = [part.strip() for part in filter_item.split(':', 1)]

                filter_obj, created = filters.objects.get_or_create(filter_title=filter_title)

                parameter_obj, created = filters_parameters.objects.get_or_create(
                    filters=filter_obj,
                    parameter_title=parameter_title
                )

                product.filters.add(parameter_obj)

    def handle_product_image(self, product, image_url, row_num):
        try:
            if not image_url:
                return

            print(f"Загрузка изображения для {product.article}: {image_url}")

            if image_url.startswith('sftp://'):
                http_url = image_url.replace('sftp://', 'https://').replace('ftpuser@94.228.120.108', 'astra-t.com')
            elif image_url.startswith('http'):
                http_url = image_url
            else:
                http_url = f'https://astra-t.com/uploads/{image_url}'

            response = requests.get(http_url, timeout=30)
            response.raise_for_status()

            if image_url.startswith(('http', 'sftp')):
                filename = os.path.basename(urlparse(image_url).path)
            else:
                filename = os.path.basename(image_url)

            product.image.save(filename, ContentFile(response.content), save=False)
            print(f"Изображение сохранено: {filename}")

        except Exception as e:
            error_msg = f"Строка {row_num}: Ошибка загрузки изображения: {str(e)}"
            self.errors.append(error_msg)
            print(error_msg)

    def process_related_data(self, product, row_data, row_num):
        self.handle_product_parameters(product, row_data, row_num)
        self.handle_product_chars(product, row_data, row_num)
        self.handle_product_slider(product, row_data, row_num)
        self.handle_product_files(product, row_data, row_num)

    def handle_product_parameters(self, product, row_data, row_num):
        from .models import product_parameters

        i = 1
        while True:
            title_key = f'product_parameter_title_{i}'
            subtitle_key = f'product_parameter_subtitle_{i}'

            title = row_data.get(title_key)
            subtitle = row_data.get(subtitle_key)

            if not title and not subtitle:
                break

            if title or subtitle:
                last_param = product_parameters.objects.filter(product=product).order_by('-order').first()
                new_order = last_param.order + 1 if last_param else 1

                param, created = product_parameters.objects.get_or_create(
                    product=product,
                    order=new_order,
                    defaults={
                        'title': title or '',
                        'subtitle': subtitle or ''
                    }
                )

                if not created:
                    param.title = title or param.title
                    param.subtitle = subtitle or param.subtitle
                    param.save()

            i += 1

    def handle_product_chars(self, product, row_data, row_num):
        from .models import product_chars

        i = 1
        while True:
            title_key = f'product_chars_title_{i}'
            subtitle_key = f'product_chars_subtitle_{i}'

            title = row_data.get(title_key)
            subtitle = row_data.get(subtitle_key)

            if not title and not subtitle:
                break

            if title or subtitle:
                last_char = product_chars.objects.filter(product=product).order_by('-order').first()
                new_order = last_char.order + 1 if last_char else 1

                char, created = product_chars.objects.get_or_create(
                    product=product,
                    order=new_order,
                    defaults={
                        'title': title or '',
                        'subtitle': subtitle or ''
                    }
                )

                if not created:
                    char.title = title or char.title
                    char.subtitle = subtitle or char.subtitle
                    char.save()

            i += 1

    def handle_product_slider(self, product, row_data, row_num):
        from .models import product_slider

        i = 1
        while True:
            image_key = f'product_slider_image_{i}'
            alt_key = f'product_slider_image_alt_{i}'

            image_url = row_data.get(image_key)
            alt_text = row_data.get(alt_key)

            if not image_url and not alt_text:
                break

            if image_url:
                try:
                    if image_url.startswith('sftp://'):
                        http_url = image_url.replace('sftp://', 'https://').replace('ftpuser@94.228.120.108', 'astra-t.com')
                    elif image_url.startswith('http'):
                        http_url = image_url
                    else:
                        http_url = f'https://astra-t.com/uploads/{image_url}'

                    response = requests.get(http_url, timeout=30)
                    response.raise_for_status()

                    if image_url.startswith(('http', 'sftp')):
                        filename = os.path.basename(urlparse(image_url).path)
                    else:
                        filename = os.path.basename(image_url)

                    last_slider = product_slider.objects.order_by('-id').first()
                    new_id = last_slider.id + 1 if last_slider else self.slider_counter
                    new_order = product_slider.objects.filter(product=product).count() + 1

                    slider = product_slider(
                        id=new_id,
                        product=product,
                        order=new_order,
                        alt=alt_text or ''
                    )

                    slider.image.save(filename, ContentFile(response.content), save=False)
                    slider.save()
                    self.slider_counter += 1

                except Exception as e:
                    self.errors.append(f"Строка {row_num}: Ошибка загрузки слайда {i}: {str(e)}")

            i += 1

    def handle_product_files(self, product, row_data, row_num):
        from .models import product_file

        i = 1
        while True:
            file_key = f'product_file_{i}'
            title_key = f'product_file_title_{i}'

            file_url = row_data.get(file_key)
            file_title = row_data.get(title_key)

            if not file_url and not file_title:
                break

            if file_url:
                try:
                    if file_url.startswith('sftp://'):
                        http_url = file_url.replace('sftp://', 'https://').replace('ftpuser@94.228.120.108', 'astra-t.com')
                    elif file_url.startswith('http'):
                        http_url = file_url
                    else:
                        http_url = f'https://astra-t.com/uploads/{file_url}'

                    response = requests.get(http_url, timeout=30)
                    response.raise_for_status()

                    if file_url.startswith(('http', 'sftp')):
                        filename = os.path.basename(urlparse(file_url).path)
                    else:
                        filename = os.path.basename(file_url)

                    last_file = product_file.objects.order_by('-id').first()
                    new_id = last_file.id + 1 if last_file else 1
                    new_order = product_file.objects.filter(product=product).count() + 1

                    file_obj = product_file(
                        id=new_id,
                        product=product,
                        order=new_order,
                        name=file_title or filename
                    )

                    file_obj.file.save(filename, ContentFile(response.content), save=False)
                    file_obj.save()

                except Exception as e:
                    self.errors.append(f"Строка {row_num}: Ошибка загрузки файла {i}: {str(e)}")

            i += 1

    def create_message(self):
        if not self.errors:
            message = f"Импорт успешно завершен. Создано: {self.success_count}, Обновлено: {self.updated_count}"
            messages.success(self.request, message)
        else:
            error_message = f"Импорт завершен с ошибками. Успешно обработано: {self.success_count + self.updated_count}. Ошибки: " + "; ".join(self.errors[:5])
            if len(self.errors) > 5:
                error_message += f" ... и еще {len(self.errors) - 5} ошибок"
            messages.error(self.request, error_message)